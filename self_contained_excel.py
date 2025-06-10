import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import json
import re
import os
import PyPDF2
import pdfplumber
from pathlib import Path
import shutil
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice
import base64
from PIL import Image as PILImage
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend

# ADD this import at the top with other imports
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    print("PIL not available - embedded Excel images won't be processed")
    PIL_AVAILABLE = False

class SelfContainedExcelFMReportGenerator:
    def __init__(self):
        self.report_data = {
            'date': '',
            'location': 'OBHUR CITY JEDDAH',
            'manpower': {},
            'cm_wo_analysis': {},
            'ppm_performance': {},
            'work_orders': [],
            'mep_tasks': {
                'electrical': [],
                'plumbing': [],
                'hvac': [],
                'fls': [],
                'civil': []
            },
            'soft_services': {
                'cleaning': [],
                'landscape': [],
                'pest_control': [],
                'waste_management': []
            }
        }
        self.raw_data = []
        self.pm_raw_data = []
        self.images_dir = "temp_images"
        os.makedirs(self.images_dir, exist_ok=True)
        self.extracted_images = {}
        self.chart_images = {}  # Store chart images as base64

    def extract_images_from_excel(self, excel_path):
        """Extract images from Excel file and store them"""
        try:
            wb = load_workbook(excel_path)
            self.extracted_images = {}
            
            # Build WO positions map
            wo_positions = {}
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in range(1, sheet.max_row + 1):
                    for col in range(1, min(20, sheet.max_column + 1)):
                        try:
                            cell_value = sheet.cell(row=row, column=col).value
                            if cell_value and str(cell_value).strip():
                                cell_str = str(cell_value).strip()
                                if cell_str.isdigit() and 6 <= len(cell_str) <= 8:
                                    if not (cell_str.startswith('202') and len(cell_str) == 8):
                                        if cell_str not in wo_positions:
                                            wo_positions[cell_str] = []
                                        wo_positions[cell_str].append((sheet_name, row, col))
                        except:
                            continue
            
            # Extract and store images as base64
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if hasattr(sheet, '_images') and sheet._images:
                    for i, image in enumerate(sheet._images):
                        try:
                            img_data = self._extract_image_data(image)
                            if img_data:
                                wo_number = self.find_wo_number_for_image(sheet, image)
                                if wo_number:
                                    image_type = self.determine_image_type(sheet, image)
                                    
                                    # Convert to base64 for embedding
                                    img_base64 = base64.b64encode(img_data).decode('utf-8')
                                    
                                    if wo_number not in self.extracted_images:
                                        self.extracted_images[wo_number] = {}
                                    self.extracted_images[wo_number][image_type] = img_base64
                                    
                        except Exception as e:
                            print(f"Error extracting image: {e}")
            
            wb.close()
            
        except Exception as e:
            print(f"Error extracting images: {e}")

    def _extract_image_data(self, image):
        """Extract image data from openpyxl image object"""
        img_data = None
        
        if hasattr(image, '_data'):
            if callable(image._data):
                try:
                    img_data = image._data()
                except Exception as e:
                    print(f"Error calling _data function: {e}")
            elif hasattr(image._data, 'getvalue'):
                image._data.seek(0)
                img_data = image._data.getvalue()
            elif hasattr(image._data, 'read'):
                image._data.seek(0)
                img_data = image._data.read()
            else:
                img_data = image._data
        
        if img_data:
            if isinstance(img_data, str):
                img_data = img_data.encode()
            elif hasattr(img_data, 'getvalue'):
                img_data = img_data.getvalue()
        
        return img_data

    def find_wo_number_for_image(self, sheet, image):
        """Find WO number associated with an image"""
        try:
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                from_cell = image.anchor._from
                if hasattr(from_cell, 'row') and hasattr(from_cell, 'col'):
                    row = from_cell.row + 1
                    col = from_cell.col + 1
                    
                    # Look for WO number in same row first
                    for check_col in range(1, col):
                        try:
                            cell_value = sheet.cell(row=row, column=check_col).value
                            if cell_value and str(cell_value).strip():
                                cell_str = str(cell_value).strip()
                                if cell_str.isdigit() and 6 <= len(cell_str) <= 8:
                                    if not (cell_str.startswith('202') and len(cell_str) == 8):
                                        return cell_str
                        except:
                            continue
                    
                    # Check nearby rows
                    for row_offset in range(1, 4):
                        for direction in [-1, 1]:
                            check_row = row + (direction * row_offset)
                            if check_row < 1:
                                continue
                            for check_col in range(1, min(10, col)):
                                try:
                                    cell_value = sheet.cell(row=check_row, column=check_col).value
                                    if cell_value and str(cell_value).strip():
                                        cell_str = str(cell_value).strip()
                                        if cell_str.isdigit() and 6 <= len(cell_str) <= 8:
                                            if not (cell_str.startswith('202') and len(cell_str) == 8):
                                                return cell_str
                                except:
                                    continue
        except Exception as e:
            print(f"Error finding WO number: {e}")
        
        return None

    def determine_image_type(self, sheet, image):
        """Determine if image is 'before' or 'after'"""
        try:
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                from_cell = image.anchor._from
                if hasattr(from_cell, 'col'):
                    col = from_cell.col
                    
                    for header_row in range(1, 5):
                        try:
                            header_value = sheet.cell(row=header_row, column=col+1).value
                            if header_value:
                                header_str = str(header_value).lower()
                                if 'before' in header_str:
                                    return 'before'
                                elif 'after' in header_str:
                                    return 'after'
                        except:
                            continue
                    
                    return 'before' if col < 10 else 'after'
        except Exception as e:
            print(f"Error determining image type: {e}")
        
        return 'before'

    def get_classification_dict(self):
        """Dictionary for classifying work orders"""
        return {
            'hvac': ['air condition'],
            'electrical': ['home appliances', 'lump', 'light', 'electrical', 'power', 'socket', 'switch', 'wiring', 'bulb', 'lamp', 'fan', 'voltage', 'circuit'],
            'plumbing': ['plumbing','water', 'sink', 'leakage', 'drain', 'toilet', 'flush', 'pipe', 'faucet', 'tap', 'shower', 'basin', 'sewage', 'blockage'],
            'fls': ['fire alarm', 'smoke detector', 'fire system', 'sprinkler', 'fire extinguisher', 'emergency light'],
            'civil': ['door', 'lock', 'civil', 'wall', 'ceiling', 'floor', 'tile', 'paint', 'window', 'glass', 'concrete']
        }

    def classify_work_order(self, description):
        """Classify work order based on description"""
        if not description or pd.isna(description):
            return 'general'
        
        description_lower = str(description).lower()
        classification_dict = self.get_classification_dict()
        
        for category, keywords in classification_dict.items():
            for keyword in keywords:
                if keyword.lower() in description_lower:
                    return category
        
        return 'general'

    def extract_date_from_filename(self, filename):
        """Extract date from filename"""
        pattern1 = r'(\d{4})-(\d{2})-(\d{2})'
        match1 = re.search(pattern1, filename)
        if match1:
            year, month, day = match1.groups()
            return f"{day}-{month}-{year}"
        
        pattern2 = r'(\d{4})_(\d{2})_(\d{2})'
        match2 = re.search(pattern2, filename)
        if match2:
            year, month, day = match2.groups()
            return f"{day}-{month}-{year}"
        
        return datetime.now().strftime('%d-%m-%Y')

    def parse_daily_report_excel(self, excel_path):
        """Parse daily report Excel file"""
        try:
            self.extract_images_from_excel(excel_path)
            df = pd.read_excel(excel_path)
            filename = Path(excel_path).name
            self.report_data['date'] = self.extract_date_from_filename(filename)
            self.raw_data = df.to_dict('records')
            self.process_work_orders()
        except Exception as e:
            print(f"Error parsing Excel file: {e}")

    def parse_pm_report_excel(self, pm_excel_path):
        """Parse PM report Excel file"""
        try:
            pm_df = pd.read_excel(pm_excel_path)
            self.pm_raw_data = pm_df.to_dict('records')
            self.calculate_ppm_from_excel()
        except Exception as e:
            print(f"Error parsing PM Excel file: {e}")

    def calculate_ppm_from_excel(self):
        """Calculate PPM performance from PM Excel data"""
        if not self.pm_raw_data:
            return
        
        valid_ppm_orders = []
        for wo in self.pm_raw_data:
            status = wo.get('status', '').upper()
            if status not in ['CAN', 'REQCAN']:
                valid_ppm_orders.append(wo)
        
        soft_service_orders = [wo for wo in valid_ppm_orders 
                            if wo.get('zzservicetype', '').upper() == 'SS']
        hard_service_orders = [wo for wo in valid_ppm_orders 
                            if wo.get('zzservicetype', '').upper() == 'HS']
        
        def count_completed(orders):
            return len([wo for wo in orders 
                    if wo.get('status', '').upper() in ['COMP', 'CLOSE']])
        
        soft_total = len(soft_service_orders)
        soft_completed = count_completed(soft_service_orders)
        hard_total = len(hard_service_orders)
        hard_completed = count_completed(hard_service_orders)
        
        self.set_ppm_data(soft_total, soft_completed, hard_total, hard_completed)

    def set_ppm_data(self, soft_total, soft_completed, hard_total, hard_completed):
        """Set PPM Performance data"""
        self.report_data['ppm_performance'] = {
            'soft_service': {
                'total': soft_total,
                'completed': soft_completed,
                'pending': soft_total - soft_completed,
                'compliance': round((soft_completed / soft_total) * 100, 1) if soft_total > 0 else 0
            },
            'hard_service': {
                'total': hard_total,
                'completed': hard_completed,
                'pending': hard_total - hard_completed,
                'compliance': round((hard_completed / hard_total) * 100, 1) if hard_total > 0 else 0
            }
        }
        
        total_tasks = soft_total + hard_total
        total_completed = soft_completed + hard_completed
        self.report_data['ppm_performance']['total'] = {
            'total': total_tasks,
            'completed': total_completed,
            'pending': total_tasks - total_completed,
            'compliance': round((total_completed / total_tasks) * 100, 1) if total_tasks > 0 else 0
        }

    def process_work_orders(self):
        """Process work orders and calculate metrics"""
        if not self.raw_data:
            return
        
        valid_work_orders = []
        for wo in self.raw_data:
            if (isinstance(wo.get('wonum'), str) and wo.get('wonum').lower() in ['wonum', 'work order', 'wo']):
                continue
            if wo.get('status', '').upper() == 'CAN':
                continue
            if not wo.get('wonum') or pd.isna(wo.get('wonum')):
                continue
            valid_work_orders.append(wo)
        
        total_raised = len(valid_work_orders)
        closed = len([wo for wo in valid_work_orders if wo.get('status', '').upper() == 'COMP'])
        pending = total_raised - closed
        
        self.set_cm_wo_data(total_raised, closed, pending)
        
        for i, wo in enumerate(valid_work_orders, 1):
            status = "RESOLVED" if wo.get('status', '').upper() == 'COMP' else "Inprogress"
            self.add_work_order(
                wo_number=str(wo.get('wonum', '')),
                sla=wo.get('wopriority', ''),
                department=wo.get('zzservicetype', ''),
                status=status
            )
        
        self.process_mep_tasks(valid_work_orders)

    def process_mep_tasks(self, work_orders):
        """Process and segregate work orders into MEP tasks"""
        for wo in work_orders:
            wo_number = str(wo.get('wonum', ''))
            description = wo.get('description', '')
            location = wo.get('location', '')
            priority = wo.get('wopriority', '')
            status = "RESOLVED" if wo.get('status', '').upper() == 'COMP' else "INPROGRESS"
            
            # Get images from extracted data
            before_image = ""
            after_image = ""
            if wo_number in self.extracted_images:
                before_image = self.extracted_images[wo_number].get('before', '')
                after_image = self.extracted_images[wo_number].get('after', '')
            
            # Extract building and apartment from location
            building_apt = ""
            if location:
                match = re.search(r'([A-Z])(\d{3})(?:-F\d{2}-(\d{3}))?', str(location))
                if match:
                    building_letter = match.group(1)
                    building_number = match.group(2)
                    apartment = match.group(3)
                    building_apt = building_letter + building_number
                    if apartment:
                        building_apt += "-" + apartment
            
            full_description = description
            if building_apt:
                full_description = f"{description} {building_apt}"
            
            category = self.classify_work_order(description)
            
            if category in self.report_data['mep_tasks']:
                self.add_mep_task(
                    category=category,
                    wo_number=wo_number,
                    skill=category.upper(),
                    description=full_description,
                    priority=priority,
                    status=status,
                    before_pic=before_image,
                    after_pic=after_image
                )

    def set_cm_wo_data(self, total_raised, closed, pending):
        """Set CM Work Order Analysis data"""
        self.report_data['cm_wo_analysis'] = {
            'total_raised': total_raised,
            'closed': closed,
            'pending': pending
        }

    def add_work_order(self, wo_number, sla, department, status, description=""):
        """Add a work order to the summary"""
        self.report_data['work_orders'].append({
            'wo_number': wo_number,
            'sla': sla,
            'department': department,
            'status': status,
            'description': description
        })

    def add_mep_task(self, category, wo_number, skill, description, priority, status, before_pic="", after_pic=""):
        """Add MEP task"""
        task = {
            'wo_number': wo_number,
            'skill': skill,
            'description': description,
            'priority': priority,
            'status': status,
            'before_pic': before_pic,
            'after_pic': after_pic
        }
        
        if category.lower() in self.report_data['mep_tasks']:
            self.report_data['mep_tasks'][category.lower()].append(task)

    def create_chart_image(self, chart_type, data, title, save_name):
        """Create chart and return as base64 string"""
        plt.figure(figsize=(10, 6))
        
        if chart_type == 'wo_analysis':
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
            
            # Pie chart
            labels = ['Closed', 'Pending']
            sizes = [data['closed'], data['pending']]
            colors = ['#2ecc71', '#e74c3c']
            
            if sum(sizes) > 0:
                ax1.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
            ax1.set_title('Work Order Status Distribution')
            
            # Bar chart
            categories = ['Total Raised', 'Closed', 'Pending']
            values = [data['total_raised'], data['closed'], data['pending']]
            bars = ax2.bar(categories, values, color=['#3498db', '#2ecc71', '#e74c3c'])
            
            ax2.set_title('Work Order Metrics')
            ax2.set_ylabel('Number of Work Orders')
            
            for bar in bars:
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2., height,
                        f'{int(height)}', ha='center', va='bottom')
        
        elif chart_type == 'ppm_performance':
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
            
            # Stacked bar chart
            categories = ['Soft Service', 'Hard Service', 'Total']
            completed = [data['soft_service']['completed'], 
                        data['hard_service']['completed'],
                        data['total']['completed']]
            pending = [data['soft_service']['pending'], 
                      data['hard_service']['pending'],
                      data['total']['pending']]
            
            ax1.bar(categories, completed, label='Completed', color='#2ecc71')
            ax1.bar(categories, pending, bottom=completed, label='Pending', color='#e74c3c')
            
            ax1.set_title('PPM Tasks Status')
            ax1.set_ylabel('Number of Tasks')
            ax1.legend()
            
            # Compliance rate chart
            services = ['Soft Service', 'Hard Service', 'Overall']
            compliance_rates = [data['soft_service']['compliance'],
                               data['hard_service']['compliance'],
                               data['total']['compliance']]
            
            bars = ax2.bar(services, compliance_rates, color=['#3498db', '#9b59b6', '#f39c12'])
            ax2.set_title('PPM Compliance Rate')
            ax2.set_ylabel('Compliance %')
            ax2.set_ylim(0, 100)
            
            for bar, rate in zip(bars, compliance_rates):
                ax2.text(bar.get_x() + bar.get_width()/2., rate + 1,
                        f'{rate}%', ha='center', va='bottom')
        
        elif chart_type == 'kpi':
            fig, ax = plt.subplots(1, 1, figsize=(12, 6))
            
            categories = ['Total PPMs Scheduled', 'Total PPMs Completed', 'PPM Compliance Rate', 'Average Time to Complete PPMs']
            values = [data['total']['total'], 
                    data['total']['completed'],
                    data['total']['compliance'],
                    24]
            
            colors = ['#2ecc71', '#2ecc71', '#f39c12', '#f39c12']
            bars = ax.bar(categories, values, color=colors)
            
            ax.set_title('Key Performance Indicator (KPI)', fontsize=16, fontweight='bold')
            ax.set_ylabel('Value')
            
            for bar, value in zip(bars, values):
                height = bar.get_height()
                if 'Compliance Rate' in categories[bars.index(bar)] or 'Average Time' in categories[bars.index(bar)]:
                    label = f'{value}%' if 'Compliance' in categories[bars.index(bar)] else f'{value} hours'
                else:
                    label = str(int(value))
                ax.text(bar.get_x() + bar.get_width()/2., height + max(values) * 0.01,
                        label, ha='center', va='bottom', fontweight='bold')
            
            plt.xticks(rotation=45, ha='right')
        
        plt.tight_layout()
        
        # Save to bytes
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
        img_buffer.seek(0)
        
        # Convert to base64
        img_base64 = base64.b64encode(img_buffer.getvalue()).decode('utf-8')
        plt.close()
        
        return img_base64

    def generate_comprehensive_excel_report(self, file_path='comprehensive_fm_report.xlsx'):
        """Generate a comprehensive Excel report with all data, charts, and images embedded"""
        
        # Create charts
        self.chart_images = {
            'wo_analysis': self.create_chart_image('wo_analysis', self.report_data['cm_wo_analysis'], 'WO Analysis', 'wo_chart'),
            'ppm_performance': self.create_chart_image('ppm_performance', self.report_data['ppm_performance'], 'PPM Performance', 'ppm_chart'),
            'kpi': self.create_chart_image('kpi', self.report_data['ppm_performance'], 'KPI Performance', 'kpi_chart')
        }
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            
            # 1. SUMMARY SHEET
            self._create_summary_sheet(writer)
            
            # 2. CM WO ANALYSIS SHEET
            self._create_wo_analysis_sheet(writer)
            
            # 3. PPM PERFORMANCE SHEET
            self._create_ppm_performance_sheet(writer)
            
            # 4. WORK ORDERS LIST SHEET
            self._create_work_orders_sheet(writer)
            
            # 5. MEP TASKS SHEETS
            self._create_mep_sheets(writer)
            
            # 6. RAW DATA SHEETS
            self._create_raw_data_sheets(writer)
        
        # Clean up temporary directory
        if os.path.exists(self.images_dir):
            shutil.rmtree(self.images_dir)
        
        return file_path

    def _create_summary_sheet(self, writer):
        """Create comprehensive summary sheet"""
        wb = writer.book
        ws = wb.create_sheet('FM_DAILY_SUMMARY', 0)
        
        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = f"{self.report_data['location']} - FM Daily Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Date: {self.report_data['date']}"
        ws['A2'].font = Font(bold=True, size=12)
        
        # CM WO Analysis Summary
        row = 4
        ws[f'A{row}'] = "CM WORK ORDER ANALYSIS"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        wo_data = [
            ['Total WOs Raised', self.report_data['cm_wo_analysis'].get('total_raised', 0)],
            ['WOs Closed', self.report_data['cm_wo_analysis'].get('closed', 0)],
            ['WOs Pending', self.report_data['cm_wo_analysis'].get('pending', 0)]
        ]
        
        for data_row in wo_data:
            row += 1
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col, value=value)
        
        # PPM Performance Summary
        row += 3
        ws[f'A{row}'] = "PPM PERFORMANCE SUMMARY"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        ppm_headers = ['Service Type', 'Total', 'Completed', 'Pending', 'Compliance %']
        for col, header in enumerate(ppm_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        ppm_data = [
            ['Soft Service', 
             self.report_data['ppm_performance'].get('soft_service', {}).get('total', 0),
             self.report_data['ppm_performance'].get('soft_service', {}).get('completed', 0),
             self.report_data['ppm_performance'].get('soft_service', {}).get('pending', 0),
             f"{self.report_data['ppm_performance'].get('soft_service', {}).get('compliance', 0)}%"],
            ['Hard Service',
             self.report_data['ppm_performance'].get('hard_service', {}).get('total', 0),
             self.report_data['ppm_performance'].get('hard_service', {}).get('completed', 0),
             self.report_data['ppm_performance'].get('hard_service', {}).get('pending', 0),
             f"{self.report_data['ppm_performance'].get('hard_service', {}).get('compliance', 0)}%"],
            ['TOTAL',
             self.report_data['ppm_performance'].get('total', {}).get('total', 0),
             self.report_data['ppm_performance'].get('total', {}).get('completed', 0),
             self.report_data['ppm_performance'].get('total', {}).get('pending', 0),
             f"{self.report_data['ppm_performance'].get('total', {}).get('compliance', 0)}%"]
        ]
        
        for data_row in ppm_data:
            row += 1
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if data_row[0] == 'TOTAL':
                    cell.font = Font(bold=True)
        
        # MEP Tasks Summary
        row += 3
        ws[f'A{row}'] = "MEP TASKS SUMMARY"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        mep_headers = ['Category', 'Total Tasks', 'Resolved Tasks']
        for col, header in enumerate(mep_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        for category, tasks in self.report_data['mep_tasks'].items():
            if tasks:
                row += 1
