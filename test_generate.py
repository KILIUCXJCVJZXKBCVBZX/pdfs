import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from docx import Document
from docx.shared import Inches
from datetime import datetime
import json
import re
import os
import PyPDF2
import pdfplumber
from pathlib import Path
import shutil
import io
from  openpyxl import load_workbook 
# ADD this import at the top with other imports
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    print("PIL not available - embedded Excel images won't be processed")
    PIL_AVAILABLE = False
from openpyxl_image_loader import SheetImageLoader

class DynamicFMReportGenerator:
    def __init__(self):
        self.report_data = {
            'date': '',
            'location': 'OBHUR CITY JEDDAH',
            'manpower': {},
            'cm_wo_analysis': {},
            'ppm_performance': {},
            'work_orders': [],
            'mep_tasks': {
                'electrical': [],
                'plumbing': [],
                'hvac': [],
                'fls': [],
                'civil': []
            },
            'soft_services': {
                'cleaning': [],
                'landscape': [],
                'pest_control': [],
                'waste_management': []
            }
        }
        self.raw_data = []
        # ADD: Separate storage for PM data
        self.pm_raw_data = []
        self.images_dir = "report_images"
        os.makedirs(self.images_dir, exist_ok=True)
        self.extracted_images = []

    def extract_images_from_excel(self, excel_path):
        """Extract images from Excel file and save them with WO-based naming"""
        try:
            # Create images directory
            images_dir = "report_images"
            os.makedirs(images_dir, exist_ok=True)
            
            wb = load_workbook(excel_path)
            self.extracted_images = {}  # Store by WO number
            
            # First, build a comprehensive map of WO numbers and their positions
            wo_positions = {}  # {wo_number: [(sheet_name, row, col), ...]}
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print(f"Scanning sheet for WO numbers: {sheet_name}")
                
                # Scan for WO numbers
                for row in range(1, sheet.max_row + 1):
                    for col in range(1, min(20, sheet.max_column + 1)):
                        try:
                            cell_value = sheet.cell(row=row, column=col).value
                            if cell_value and str(cell_value).strip():
                                cell_str = str(cell_value).strip()
                                if cell_str.isdigit() and 6 <= len(cell_str) <= 8:
                                    if not (cell_str.startswith('202') and len(cell_str) == 8):
                                        if cell_str not in wo_positions:
                                            wo_positions[cell_str] = []
                                        wo_positions[cell_str].append((sheet_name, row, col))
                        except:
                            continue
            
            print(f"Found WO positions: {wo_positions}")
            
            # Process images and save with predictable naming
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                if hasattr(sheet, '_images') and sheet._images:
                    print(f"Found {len(sheet._images)} images in sheet: {sheet_name}")
                    
                    for i, image in enumerate(sheet._images):
                        try:
                            # Extract image data (keep existing logic)
                            img_data = None
                            
                            if hasattr(image, '_data'):
                                if callable(image._data):
                                    try:
                                        img_data = image._data()
                                    except Exception as e:
                                        print(f"Error calling _data function: {e}")
                                elif hasattr(image._data, 'getvalue'):
                                    image._data.seek(0)
                                    img_data = image._data.getvalue()
                                elif hasattr(image._data, 'read'):
                                    image._data.seek(0)
                                    img_data = image._data.read()
                                else:
                                    img_data = image._data
                            
                            if img_data:
                                if isinstance(img_data, str):
                                    img_data = img_data.encode()
                                elif hasattr(img_data, 'getvalue'):
                                    img_data = img_data.getvalue()
                                
                                # Find the closest WO number
                                wo_number = self.find_wo_number_for_image(sheet, image)
                                
                                if wo_number:
                                    # Determine image type (before/after)
                                    image_type = self.determine_image_type(sheet, image)
                                    
                                    # **CHANGE: Create predictable filename**
                                    filename = f"{wo_number}_{image_type}.png"
                                    filepath = os.path.join(images_dir, filename)
                                    
                                    # Save the image
                                    with open(filepath, 'wb') as f:
                                        f.write(img_data)
                                    print(f"Saved image: {filepath}")
                                    
                                    # **CHANGE: Store in simplified structure**
                                    if wo_number not in self.extracted_images:
                                        self.extracted_images[wo_number] = {}
                                    self.extracted_images[wo_number][image_type] = filepath
                                    
                                    print(f"Mapped image to WO {wo_number} as {image_type}: {filepath}")
                                else:
                                    # Save with generic name for unmapped images
                                    filename = f"unknown_{sheet_name}_{i+1}.png"
                                    filepath = os.path.join(images_dir, filename)
                                    with open(filepath, 'wb') as f:
                                        f.write(img_data)
                                    print(f"Saved unmapped image: {filepath}")
                                    
                        except Exception as e:
                            print(f"Error extracting individual image: {e}")
            
            wb.close()
            print(f"Final extracted images mapping: {self.extracted_images}")
            
        except Exception as e:
            print(f"Error extracting images: {e}")

    def find_wo_number_for_image(self, sheet, image):
        """Try to find the WO number associated with an image with improved logic"""
        try:
            # Get image position
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                from_cell = image.anchor._from
                if hasattr(from_cell, 'row') and hasattr(from_cell, 'col'):
                    row = from_cell.row + 1  # openpyxl is 0-indexed, Excel is 1-indexed
                    col = from_cell.col + 1
                    
                    print(f"Image found at row {row}, col {col}")
                    
                    # Strategy 1: Look for WO number in the same row first (most likely scenario)
                    # Check columns to the left of the image (WO numbers are usually in earlier columns)
                    for check_col in range(1, col):  # Check from column A to the image column
                        try:
                            cell_value = sheet.cell(row=row, column=check_col).value
                            if cell_value and str(cell_value).strip():
                                cell_str = str(cell_value).strip()
                                # Check if this looks like a WO number (digits, length 6-8)
                                if cell_str.isdigit() and 6 <= len(cell_str) <= 8:
                                    print(f"Found WO number {cell_str} in same row {row}, col {check_col}")
                                    return cell_str
                        except:
                            continue
                    
                    # Strategy 2: If not found in same row, check nearby rows but prioritize closer rows
                    # Check rows above and below, but give preference to exact row matches
                    for row_offset in range(1, 4):  # Check up to 3 rows above/below
                        for direction in [-1, 1]:  # Check above first, then below
                            check_row = row + (direction * row_offset)
                            if check_row < 1:  # Skip invalid rows
                                continue
                                
                            # Look for WO number in the first few columns of this row
                            for check_col in range(1, min(10, col)):  # Check first 9 columns or up to image column
                                try:
                                    cell_value = sheet.cell(row=check_row, column=check_col).value
                                    if cell_value and str(cell_value).strip():
                                        cell_str = str(cell_value).strip()
                                        # Check if this looks like a WO number (digits, length 6-8)
                                        if cell_str.isdigit() and 6 <= len(cell_str) <= 8:
                                            print(f"Found WO number {cell_str} at nearby row {check_row}, col {check_col} (offset: {direction * row_offset})")
                                            return cell_str
                                except:
                                    continue
                    
                    # Strategy 3: Look for WO patterns with regex in a more targeted way
                    # Focus on the same row and immediate vicinity
                    for check_row in range(max(1, row-2), row+3):
                        for check_col in range(1, min(15, col+3)):  # Reasonable column range
                            try:
                                cell_value = sheet.cell(row=check_row, column=check_col).value
                                if cell_value:
                                    cell_str = str(cell_value).strip()
                                    # Look for patterns like WO followed by numbers or just numbers
                                    import re
                                    # Try to find 6-8 digit numbers
                                    wo_match = re.search(r'(\d{6,8})', cell_str)
                                    if wo_match:
                                        wo_num = wo_match.group(1)
                                        # Additional validation: avoid dates (like 20250607)
                                        if not (wo_num.startswith('202') and len(wo_num) == 8):
                                            print(f"Found WO pattern {wo_num} in cell {cell_str} at row {check_row}, col {check_col}")
                                            return wo_num
                            except:
                                continue
                    
                    # Strategy 4: If still not found, try to find based on column headers
                    # Look for a column header that indicates WO numbers
                    wo_column = None
                    for check_col in range(1, 20):  # Check first 20 columns
                        try:
                            # Check first few rows for headers
                            for header_row in range(1, 5):
                                header_value = sheet.cell(row=header_row, column=check_col).value
                                if header_value:
                                    header_str = str(header_value).lower().strip()
                                    if 'wonum' in header_str or 'wo' in header_str or 'work order' in header_str:
                                        wo_column = check_col
                                        print(f"Found WO column {wo_column} with header '{header_value}'")
                                        break
                            if wo_column:
                                break
                        except:
                            continue
                    
                    # If we found a WO column, look for the WO number in that column for our row
                    if wo_column:
                        for check_row in range(max(1, row-2), row+3):  # Check nearby rows
                            try:
                                cell_value = sheet.cell(row=check_row, column=wo_column).value
                                if cell_value and str(cell_value).strip():
                                    cell_str = str(cell_value).strip()
                                    if cell_str.isdigit() and 6 <= len(cell_str) <= 8:
                                        print(f"Found WO number {cell_str} in WO column {wo_column}, row {check_row}")
                                        return cell_str
                            except:
                                continue
                                
        except Exception as e:
            print(f"Error finding WO number for image: {e}")
        
        print(f"Could not find WO number for image at row {row if 'row' in locals() else 'unknown'}")
        return None


    def determine_image_type(self, sheet, image):
        """Determine if image is 'before' or 'after' based on column position"""
        try:
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                from_cell = image.anchor._from
                if hasattr(from_cell, 'col'):
                    col = from_cell.col
                    
                    # Look at column headers to determine type
                    for header_row in range(1, 5):  # Check first few rows for headers
                        try:
                            header_value = sheet.cell(row=header_row, column=col+1).value
                            if header_value:
                                header_str = str(header_value).lower()
                                if 'before' in header_str:
                                    return 'before'
                                elif 'after' in header_str:
                                    return 'after'
                        except:
                            continue
                    
                    # Fallback: assume earlier columns are 'before', later are 'after'
                    if col < 10:  # Arbitrary threshold
                        return 'before'
                    else:
                        return 'after'
        except Exception as e:
            print(f"Error determining image type: {e}")
        
        return 'before'  # Default fallback

    def process_image_path(self, image_path, wo_number=None, image_type=None):
        """Process and validate image paths with predictable WO-based naming"""
        print(f"Processing image path for WO {wo_number}, type {image_type}")
        
        # **CHANGE: First try predictable path based on WO number**
        if wo_number and image_type:
            # Try the predictable filename pattern
            predictable_filename = f"{wo_number}_{image_type}.png"
            predictable_path = os.path.join(self.images_dir, predictable_filename)
            
            if os.path.exists(predictable_path):
                print(f"Found image using predictable naming: {predictable_path}")
                return predictable_path
        
        # **CHANGE: Then check extracted images mapping**
        if wo_number and str(wo_number) in self.extracted_images:
            print(f"Found extracted images for WO {wo_number}: {self.extracted_images[str(wo_number)]}")
            if image_type and image_type in self.extracted_images[str(wo_number)]:
                extracted_path = self.extracted_images[str(wo_number)][image_type]
                if os.path.exists(extracted_path):
                    print(f"Using extracted image for WO {wo_number} ({image_type}): {extracted_path}")
                    return extracted_path
        
        # Handle Excel data (NaN/empty values)
        if not image_path or pd.isna(image_path):
            print(f"No image path provided for WO {wo_number} ({image_type})")
            return ""
        
        # Handle numeric values (Excel sometimes stores as float)
        if isinstance(image_path, (int, float)):
            print(f"Image path is numeric (NaN/float) for WO {wo_number}")
            return ""
        
        # Rest of the existing logic for other path types...
        image_path = str(image_path).strip()
        if not image_path:
            return ""
        
        # Check if it's a URL - return as is
        if image_path.startswith(('http://', 'https://')):
            return image_path
        
        # Check if it's base64 data - return as is
        if image_path.startswith('data:image'):
            return image_path
        
        # Check if it's a local file path
        if os.path.exists(image_path):
            return image_path
        
        # Try to find image in common directories (existing logic)
        possible_paths = [
            image_path,
            os.path.join('images', os.path.basename(image_path)),
            os.path.join('assets', os.path.basename(image_path)),
            os.path.join(self.images_dir, os.path.basename(image_path)),
            os.path.basename(image_path)
        ]
        
        for path in possible_paths:
            if os.path.exists(path):
                print(f"Found image at: {path}")
                return path
        
        print(f"Image not found: {image_path}")
        return ""

    
    def copy_image_to_report_dir(self, image_path, wo_number, image_type):
        """Copy image to report directory with WO-based naming (simplified)"""
        if not image_path:
            return ""
        
        # Don't copy URLs or base64 data
        if isinstance(image_path, str) and image_path.startswith(('http://', 'https://', 'data:image')):
            return image_path
        
        # **CHANGE: Check if image is already in correct location with correct name**
        expected_filename = f"{wo_number}_{image_type}.png"
        expected_path = os.path.join(self.images_dir, expected_filename)
        
        # If image is already where we expect it, return that path
        if os.path.exists(expected_path):
            print(f"Image already exists in correct location: {expected_path}")
            return expected_path
        
        # If image exists elsewhere, copy it to the expected location
        if os.path.exists(str(image_path)):
            try:
                shutil.copy2(image_path, expected_path)
                print(f"Copied image from {image_path} to {expected_path}")
                return expected_path
            except Exception as e:
                print(f"Error copying image {image_path}: {e}")
                return ""
        
        print(f"Image file not found: {image_path}")
        return ""
        
    def get_classification_dict(self):
        """Dictionary for classifying work orders by description keywords"""
        return {'hvac': ['air condition'],
            'electrical': ['home appliances', 'lump', 'light', 'electrical', 'power', 'socket', 'switch', 'wiring', 'bulb', 'lamp', 'fan', 'voltage', 'circuit'],
            'plumbing': ['plumbing','water', 'sink', 'leakage', 'drain', 'toilet', 'flush', 'pipe', 'faucet', 'tap', 'shower', 'basin', 'sewage', 'blockage'],
            'fls': ['fire alarm', 'smoke detector', 'fire system', 'sprinkler', 'fire extinguisher', 'emergency light'],
            'civil': ['door', 'lock', 'civil', 'wall', 'ceiling', 'floor', 'tile', 'paint', 'window', 'glass', 'concrete']
        }

    def classify_work_order(self, description):
        """Classify work order based on description keywords"""
        if not description or pd.isna(description):
            return 'general'
        
        description_lower = str(description).lower()
        classification_dict = self.get_classification_dict()
        
        for category, keywords in classification_dict.items():
            for keyword in keywords:
                if keyword.lower() in description_lower:
                    return category
        
        return 'general'
    
    def extract_date_from_filename(self, filename):
        """Extract date from filename like daily_report_2025-06-07.xlsx or Daily_job_listed_Report_2025_06_07.pdf"""
        # Pattern for YYYY-MM-DD format
        pattern1 = r'(\d{4})-(\d{2})-(\d{2})'
        match1 = re.search(pattern1, filename)
        if match1:
            year, month, day = match1.groups()
            return f"{day}-{month}-{year}"
        
        # Pattern for YYYY_MM_DD format  
        pattern2 = r'(\d{4})_(\d{2})_(\d{2})'
        match2 = re.search(pattern2, filename)
        if match2:
            year, month, day = match2.groups()
            return f"{day}-{month}-{year}"
        
        return datetime.now().strftime('%d-%m-%Y')
    
    def parse_daily_report_excel(self, excel_path):
        """Parse the daily report Excel file and extract work order data"""
        try:
            # First extract embedded images
            self.extract_images_from_excel(excel_path)
            
            # Read the Excel file
            df = pd.read_excel(excel_path)
            
            # Extract date from filename
            filename = Path(excel_path).name
            self.report_data['date'] = self.extract_date_from_filename(filename)
            
            # Convert DataFrame to list of dictionaries for processing
            self.raw_data = df.to_dict('records')
            
            # Process the work orders
            self.process_work_orders()

        except Exception as e:
            print(f"Error parsing Excel file: {e}")
    def parse_pm_report_excel(self, pm_excel_path):
        """Parse the PM report Excel file and extract PPM data"""
        try:
            # Read the PM Excel file
            pm_df = pd.read_excel(pm_excel_path)
            
            # Convert DataFrame to list of dictionaries for processing
            self.pm_raw_data = pm_df.to_dict('records')
            
            # Process the PPM data
            self.calculate_ppm_from_excel()
            
        except Exception as e:
            print(f"Error parsing PM Excel file: {e}")
    def calculate_ppm_from_excel(self):
        """Calculate PPM performance from PM Excel data"""
        if not self.pm_raw_data:
            return
        
        # Filter out cancelled work orders (CAN, REQCAN)
        valid_ppm_orders = []
        for wo in self.pm_raw_data:
            status = wo.get('status', '').upper()
            if status not in ['CAN', 'REQCAN']:
                valid_ppm_orders.append(wo)
        
        # Segregate by service type
        soft_service_orders = [wo for wo in valid_ppm_orders 
                            if wo.get('zzservicetype', '').upper() == 'SS']
        hard_service_orders = [wo for wo in valid_ppm_orders 
                            if wo.get('zzservicetype', '').upper() == 'HS']
        
        # Calculate completed (COMP, CLOSE status)
        def count_completed(orders):
            return len([wo for wo in orders 
                    if wo.get('status', '').upper() in ['COMP', 'CLOSE']])
        
        # Calculate metrics
        soft_total = len(soft_service_orders)
        soft_completed = count_completed(soft_service_orders)
        
        hard_total = len(hard_service_orders)
        hard_completed = count_completed(hard_service_orders)
        
        # Set the PPM data
        self.set_ppm_data(soft_total, soft_completed, hard_total, hard_completed)
    def set_ppm_data(self, soft_total, soft_completed, hard_total, hard_completed):
        """Set PPM Performance data"""
        self.report_data['ppm_performance'] = {
            'soft_service': {
                'total': soft_total,
                'completed': soft_completed,
                'pending': soft_total - soft_completed,
                'compliance': round((soft_completed / soft_total) * 100, 1) if soft_total > 0 else 0
            },
            'hard_service': {
                'total': hard_total,
                'completed': hard_completed,
                'pending': hard_total - hard_completed,
                'compliance': round((hard_completed / hard_total) * 100, 1) if hard_total > 0 else 0
            }
        }
        
        # Calculate totals
        total_tasks = soft_total + hard_total
        total_completed = soft_completed + hard_completed
        self.report_data['ppm_performance']['total'] = {
            'total': total_tasks,
            'completed': total_completed,
            'pending': total_tasks - total_completed,
            'compliance': round((total_completed / total_tasks) * 100, 1) if total_tasks > 0 else 0
        }

    # 3. MODIFY: Process work orders method to handle dynamic data
    def process_work_orders(self):
        """Process work orders and calculate metrics"""
        if not self.raw_data:
            return
        
        # Filter out cancelled work orders and any header rows
        valid_work_orders = []
        for wo in self.raw_data:
            # Skip if it's a header row or invalid data
            if (isinstance(wo.get('wonum'), str) and wo.get('wonum').lower() in ['wonum', 'work order', 'wo']):
                continue
            # Skip cancelled work orders
            if wo.get('status', '').upper() == 'CAN':
                continue
            # Skip rows with missing essential data
            if not wo.get('wonum') or pd.isna(wo.get('wonum')):
                continue
                
            valid_work_orders.append(wo)
        
        # Calculate CM Work Order Analysis
        total_raised = len(valid_work_orders)
        closed = len([wo for wo in valid_work_orders if wo.get('status', '').upper() == 'COMP'])
        pending = total_raised - closed
        
        self.set_cm_wo_data(total_raised, closed, pending)
        
        # Process work orders for summary
        for i, wo in enumerate(valid_work_orders, 1):
            status = "RESOLVED" if wo.get('status', '').upper() == 'COMP' else "Inprogress"
            self.add_work_order(
                wo_number=str(wo.get('wonum', '')),
                sla=wo.get('wopriority', ''),
                department=wo.get('zzservicetype', ''),
                status=status
            )
        self.process_mep_tasks(valid_work_orders)


    def process_mep_tasks(self, work_orders):
        """Process and segregate work orders into MEP tasks"""
        for wo in work_orders:
            # Get work order details
            wo_number = str(wo.get('wonum', ''))
            description = wo.get('description', '')
            location = wo.get('location', '')
            priority = wo.get('wopriority', '')
            status = "RESOLVED" if wo.get('status', '').upper() == 'COMP' else "INPROGRESS"
            
            # Get image paths from Excel
            before_image_raw = wo.get('Before_Image', '')
            after_image_raw = wo.get('After_Image', '')

            print(f"Processing WO {wo_number}:")
            print(f"  Before image raw: {before_image_raw} (type: {type(before_image_raw)})")
            print(f"  After image raw: {after_image_raw} (type: {type(after_image_raw)})")

            # Process and validate image paths
            before_image = self.process_image_path(before_image_raw, wo_number, 'before')
            after_image = self.process_image_path(after_image_raw, wo_number, 'after')

            print(f"  Processed before image: {before_image}")
            print(f"  Processed after image: {after_image}")

            # Copy images to report directory if they exist
            if before_image:
                before_image = self.copy_image_to_report_dir(before_image, wo_number, 'before')
                print(f"  Final before image path: {before_image}")
            if after_image:
                after_image = self.copy_image_to_report_dir(after_image, wo_number, 'after')
                print(f"  Final after image path: {after_image}")
            
            # Extract building and apartment from location (e.g., "842-301")
            building_apt = ""
            if location:
                match = re.search(r'([A-Z])(\d{3})(?:-F\d{2}-(\d{3}))?', str(location))
                if match:
                    building_letter = match.group(1)  # A, B, C, etc.
                    building_number = match.group(2)  # 813, 816, etc.
                    apartment = match.group(3)        # 602 or None
                    building_apt = building_letter + building_number
                    if apartment:
                        building_apt += "-" + apartment
            
            # Combine description with building/apartment info
            full_description = description
            if building_apt:
                full_description = f"{description} {building_apt}"
            
            # Classify the work order
            category = self.classify_work_order(description)
            
            # Add to appropriate MEP category
            if category in self.report_data['mep_tasks']:
                self.add_mep_task(
                    category=category,
                    wo_number=wo_number,
                    skill=category.upper(),
                    description=full_description,
                    priority=priority,
                    status=status,
                    before_pic=before_image,
                    after_pic=after_image
                )
                print(f"  Added to {category} category")
            print("  ---")
    
    def set_cm_wo_data(self, total_raised, closed, pending):
        """Set CM Work Order Analysis data"""
        self.report_data['cm_wo_analysis'] = {
            'total_raised': total_raised,
            'closed': closed,
            'pending': pending
        }
    
    def add_work_order(self, wo_number, sla, department, status, description=""):
        """Add a work order to the summary"""
        self.report_data['work_orders'].append({
            'wo_number': wo_number,
            'sla': sla,
            'department': department,
            'status': status,
            'description': description
        })
    def generate_kpi_chart(self, save_path='kpi_chart.png'):
        """Generate KPI Performance chart"""
        ppm_data = self.report_data['ppm_performance']
        
        fig, ax = plt.subplots(1, 1, figsize=(12, 6))
        
        # KPI metrics
        categories = ['Total PPMs Scheduled', 'Total PPMs Completed', 'PPM Compliance Rate', 'Average Time to Complete PPMs']
        values = [ppm_data['total']['total'], 
                ppm_data['total']['completed'],
                ppm_data['total']['compliance'],
                24]  # Assuming 24 hours as shown in your image
        
        # Create bars with different colors
        colors = ['#2ecc71', '#2ecc71', '#f39c12', '#f39c12']
        bars = ax.bar(categories, values, color=colors)
        
        ax.set_title('Key Performance Indicator (KPI)', fontsize=16, fontweight='bold')
        ax.set_ylabel('Value')
        
        # Add value labels on bars
        for bar, value in zip(bars, values):
            height = bar.get_height()
            if 'Compliance Rate' in categories[bars.index(bar)] or 'Average Time' in categories[bars.index(bar)]:
                label = f'{value}%' if 'Compliance' in categories[bars.index(bar)] else f'{value} hours'
            else:
                label = str(int(value))
            ax.text(bar.get_x() + bar.get_width()/2., height + max(values) * 0.01,
                    label, ha='center', va='bottom', fontweight='bold')
        
        # Rotate x-axis labels for better readability
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return save_path
    
    def add_mep_task(self, category, wo_number, skill, description, priority, status, before_pic="", after_pic=""):
        """Add MEP task (electrical, plumbing, hvac, fls, civil)"""
        task = {
            'wo_number': wo_number,
            'skill': skill,
            'description': description,
            'priority': priority,
            'status': status,
            'before_pic': before_pic,
            'after_pic': after_pic
        }
        
        if category.lower() in self.report_data['mep_tasks']:
            self.report_data['mep_tasks'][category.lower()].append(task)
        
    def get_image_html(self, image_path, alt_text="Image"):
        """Generate HTML for image display"""
        if not image_path:
            return '<div class="image-container"><span class="image-error">No Image</span></div>'
        
        # Handle different image path types
        if image_path.startswith(('http://', 'https://')):
            # URL - use as is
            src = image_path
        elif image_path.startswith('data:image'):
            # Base64 data - use as is
            src = image_path
        elif os.path.exists(image_path):
            # Local file - use relative path and ensure forward slashes for web
            src = os.path.relpath(image_path).replace('\\', '/')
            print(f"Using image source: {src}")
        else:
            # File not found
            print(f"Image not found for HTML: {image_path}")
            return f'<div class="image-container"><span class="image-error">Image not found:<br>{os.path.basename(image_path) if image_path else "N/A"}</span></div>'
        
        return f'''
        <div class="image-container">
            <img src="{src}" 
                style="width:100px;height:100px;object-fit:cover;border:1px solid #ddd;" 
                alt="{alt_text}" 
                onerror="this.style.display='none'; this.parentNode.querySelector('.image-error').style.display='block';" />
            <span class="image-error" style="display:none;">Image Error</span>
        </div>
        '''
    
    def set_manpower(self, manpower_dict):
        """Set manpower data"""
        self.report_data['manpower'] = manpower_dict
    
    def generate_wo_analysis_chart(self, save_path='wo_analysis_chart.png'):
        """Generate CM Work Order Analysis chart"""
        data = self.report_data['cm_wo_analysis']
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        
        # Pie chart for WO status
        labels = ['Closed', 'Pending']
        sizes = [data['closed'], data['pending']]
        colors = ['#2ecc71', '#e74c3c']
        
        # Only create pie chart if there's data
        if sum(sizes) > 0:
            ax1.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        ax1.set_title('Work Order Status Distribution')
        
        # Bar chart for WO metrics
        categories = ['Total Raised', 'Closed', 'Pending']
        values = [data['total_raised'], data['closed'], data['pending']]
        bars = ax2.bar(categories, values, color=['#3498db', '#2ecc71', '#e74c3c'])
        
        ax2.set_title('Work Order Metrics')
        ax2.set_ylabel('Number of Work Orders')
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            ax2.text(bar.get_x() + bar.get_width()/2., height,
                    f'{int(height)}', ha='center', va='bottom')
        
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return save_path
    
    def generate_ppm_chart(self, save_path='ppm_chart.png'):
        """Generate PPM Performance chart"""
        ppm_data = self.report_data['ppm_performance']
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
        
        # Stacked bar chart for PPM tasks
        categories = ['Soft Service', 'Hard Service', 'Total']
        completed = [ppm_data['soft_service']['completed'], 
                    ppm_data['hard_service']['completed'],
                    ppm_data['total']['completed']]
        pending = [ppm_data['soft_service']['pending'], 
                  ppm_data['hard_service']['pending'],
                  ppm_data['total']['pending']]
        
        ax1.bar(categories, completed, label='Completed', color='#2ecc71')
        ax1.bar(categories, pending, bottom=completed, label='Pending', color='#e74c3c')
        
        ax1.set_title('PPM Tasks Status')
        ax1.set_ylabel('Number of Tasks')
        ax1.legend()
        
        # Compliance rate chart
        services = ['Soft Service', 'Hard Service', 'Overall']
        compliance_rates = [ppm_data['soft_service']['compliance'],
                           ppm_data['hard_service']['compliance'],
                           ppm_data['total']['compliance']]
        
        bars = ax2.bar(services, compliance_rates, color=['#3498db', '#9b59b6', '#f39c12'])
        ax2.set_ylabel('Compliance %')
        ax2.set_ylim(0, 100)
        
        # Add percentage labels
        for bar, rate in zip(bars, compliance_rates):
            ax2.text(bar.get_x() + bar.get_width()/2., rate + 1,
                    f'{rate}%', ha='center', va='bottom')
        
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return save_path
    
    def generate_html_report(self, template_path='fm_report_template.html'):
        """Generate HTML report template"""
        html_template = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>FM Daily Report - {self.report_data['date']}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .section {{ margin: 20px 0; }}
                table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .status-resolved {{ color: green; font-weight: bold; }}
                .status-pending {{ color: red; font-weight: bold; }}
                .priority-1 {{ background-color: #ffebee; }}
                .priority-2 {{ background-color: #fff3e0; }}
                .priority-3 {{ background-color: #f3e5f5; }}
                .chart-container {{ margin: 20px 0; text-align: center; }}
                .mep-table img {{
                    max-width: 100%;
                    max-height: 100%;
                    object-fit: cover;
                }}
                .mep-table td {{
                    vertical-align: middle;
                    text-align: center;
                }}
                .image-container {{
                    width: 100px;
                    height: 100px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    border: 1px solid #ddd;
                    background-color: #f9f9f9;
                    margin: 2px auto;
                }}


                .image-error {{
                        color: #999;
                        font-size: 11px;
                        text-align: center;
                        padding: 5px;
                }}
                .mep-table td {{
                vertical-align: middle;
                text-align: center;
                padding: 5px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{self.report_data['location']}</h1>
                <h2>FM Daily Report</h2>
                <h3>{self.report_data['date']}</h3>
            </div>
            
            <div class="section">
                <h2>3. CM Work Order (WO) Analysis & Preventive Maintenance (PPM) Performance Dashboard</h2>
                
                <h3>a) CM Work Order (WO) Analysis</h3>
                <table>
                    <tr><th>Metric</th><th>Value</th></tr>
                    <tr><td>Total Number of WOs Raised</td><td>{self.report_data['cm_wo_analysis'].get('total_raised', 0)}</td></tr>
                    <tr><td>Number of WOs Closed</td><td>{self.report_data['cm_wo_analysis'].get('closed', 0)}</td></tr>
                    <tr><td>Number of WOs Currently Pending</td><td>{self.report_data['cm_wo_analysis'].get('pending', 0)}</td></tr>
                </table>
                
                <div class="chart-container">
                    <img src="wo_analysis_chart.png" alt="WO Analysis Chart" style="max-width: 100%;">
                </div>
                
                <h3>b) Preventive Maintenance (PPM) Performance</h3>
                <table>
                    <tr><th>Services Description</th><th>Total PPM Tasks</th><th>Completed</th><th>Pending</th><th>Compliance %</th></tr>
                    <tr>
                        <td>Soft Service</td>
                        <td>{self.report_data['ppm_performance'].get('soft_service', {}).get('total', 0)}</td>
                        <td>{self.report_data['ppm_performance'].get('soft_service', {}).get('completed', 0)}</td>
                        <td>{self.report_data['ppm_performance'].get('soft_service', {}).get('pending', 0)}</td>
                        <td>{self.report_data['ppm_performance'].get('soft_service', {}).get('compliance', 0)}%</td>
                    </tr>
                    <tr>
                        <td>Hard Service</td>
                        <td>{self.report_data['ppm_performance'].get('hard_service', {}).get('total', 0)}</td>
                        <td>{self.report_data['ppm_performance'].get('hard_service', {}).get('completed', 0)}</td>
                        <td>{self.report_data['ppm_performance'].get('hard_service', {}).get('pending', 0)}</td>
                        <td>{self.report_data['ppm_performance'].get('hard_service', {}).get('compliance', 0)}%</td>
                    </tr>
                    <tr style="font-weight: bold;">
                        <td>Total</td>
                        <td>{self.report_data['ppm_performance'].get('total', {}).get('total', 0)}</td>
                        <td>{self.report_data['ppm_performance'].get('total', {}).get('completed', 0)}</td>
                        <td>{self.report_data['ppm_performance'].get('total', {}).get('pending', 0)}</td>
                        <td>{self.report_data['ppm_performance'].get('total', {}).get('compliance', 0)}%</td>
                    </tr>
                </table>
                
                <div class="chart-container">
                    <img src="ppm_chart.png" alt="PPM Performance Chart" style="max-width: 100%;">
                </div>
                                <h3>c) Key Performance Indicator (KPI)</h3>
                <table>
                    <tr><th>KPI</th><th>Value</th></tr>
                    <tr><td>Total PPMs Scheduled</td><td>{self.report_data['ppm_performance'].get('total', {}).get('total', 0)}</td></tr>
                    <tr><td>Total PPMs Completed</td><td>{self.report_data['ppm_performance'].get('total', {}).get('completed', 0)}</td></tr>
                    <tr><td>PPM Compliance Rate</td><td>{self.report_data['ppm_performance'].get('total', {}).get('compliance', 0)}%</td></tr>
                    <tr><td>Average Time to Complete PPMs</td><td>24 hours</td></tr>
                </table>
                
                <div class="chart-container">
                    <img src="kpi_chart.png" alt="KPI Performance Chart" style="max-width: 100%;">
                </div>

            </div>
            
            <div class="section">
                <h2>Maximo Work Order Summary</h2>
                <table>
                    <tr><th>S.N</th><th>WO</th><th>SLA's</th><th>Department</th><th>Remarks</th></tr>
            
        """
        # CHANGE: Sort work orders to show RESOLVED first, then INPROGRESS
        sorted_work_orders = sorted(self.report_data['work_orders'], 
            key=lambda x: (0 if x['status'].lower() == 'resolved' else 1, x['wo_number']))
        # Add work orders
        for i, wo in enumerate(sorted_work_orders, 1):
            status_class = 'status-resolved' if wo['status'].lower() == 'resolved' else 'status-pending'
            html_template += f"""
                    <tr>
                        <td>{i}</td>
                        <td>{wo['wo_number']}</td>
                        <td>{wo['sla']}</td>
                        <td>{wo['department']}</td>
                        <td class="{status_class}">{wo['status']}</td>
                    </tr>
            """
        
        html_template += """
                </table>
        
            </div>
        """
        for category, tasks in self.report_data['mep_tasks'].items():
            if tasks:  # Only show categories that have tasks
                category_title = category.replace('_', ' ').title()
                if category.lower() == 'fls':
                    category_title = 'FLS'
                elif category.lower() == 'hvac':
                    category_title = 'HVAC'
                sorted_tasks = sorted(tasks, key=lambda x: (0 if x['status'].lower() == 'resolved' else 1, x['wo_number']))
                html_template += f"""
                    <div class="section">
                        <h2>{category_title}</h2>
                        <table>
                            <tr>
                                <th>SN.</th>
                                <th>Skill</th>
                                <th>W/O</th>
                                <th>Description</th>
                                <th>Priority</th>
                                <th>Before Pic</th>
                                <th>After Pic</th>
                                <th>Status</th>
                            </tr>
                """
                
                for i, task in enumerate(sorted_tasks, 1):
                    status_class = 'status-resolved' if task['status'].lower() == 'resolved' else 'status-pending'
                    before_img = self.get_image_html(task['before_pic'], 'Before')
                    after_img = self.get_image_html(task['after_pic'], 'After')
                    
                    html_template += f"""
                            <tr>
                                <td>{i}</td>
                                <td>{task['skill']}</td>
                                <td>{task['wo_number']}</td>
                                <td>{task['description']}</td>
                                <td>{task['priority']}</td>
                                <td>{before_img}</td>
                                <td>{after_img}</td>
                                <td class="{status_class}">{task['status']}</td>
                            </tr>
                    """
                
                html_template += """
                        </table>
                        </div>
                    </body>
                </html>
                """
        
        with open(template_path, 'w', encoding='utf-8') as f:
            f.write(html_template)
        
        return template_path
    
    def generate_excel_report(self, file_path='fm_report.xlsx'):
        """Generate Excel report with multiple sheets"""
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # CM WO Analysis sheet
            wo_df = pd.DataFrame([self.report_data['cm_wo_analysis']])
            wo_df.to_excel(writer, sheet_name='CM_WO_Analysis', index=False)
            
            # PPM Performance sheet
            ppm_data = []
            for service_type, data in self.report_data['ppm_performance'].items():
                if isinstance(data, dict):
                    row = {'Service_Type': service_type.replace('_', ' ').title()}
                    row.update(data)
                    ppm_data.append(row)
            
            ppm_df = pd.DataFrame(ppm_data)
            ppm_df.to_excel(writer, sheet_name='PPM_Performance', index=False)
            
            # Work Orders sheet
            wo_list_df = pd.DataFrame(self.report_data['work_orders'])
            wo_list_df.to_excel(writer, sheet_name='Work_Orders', index=False)
            
            # MEP Tasks sheets
            for category, tasks in self.report_data['mep_tasks'].items():
                if tasks:
                    mep_df = pd.DataFrame(tasks)
                    sheet_name = f'MEP_{category.title()}'
                    mep_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return file_path

# 5. MODIFY: The main usage function
def dynamic_example_usage(excel_file_path,pm_excel_file_path):
    """Example of how to use the Dynamic FM Report Generator with Excel input"""
    
    # Create generator instance
    fm_report = DynamicFMReportGenerator()
    
    # Parse the daily report Excel file
    fm_report.parse_daily_report_excel(excel_file_path)

    
    
    # Parse the PM Excel file
    fm_report.parse_pm_report_excel(pm_excel_file_path)
    
    # Generate reports
    fm_report.generate_wo_analysis_chart()
    fm_report.generate_ppm_chart()
    fm_report.generate_kpi_chart()
    fm_report.generate_html_report()
    fm_report.generate_excel_report()
    for category, tasks in fm_report.report_data['mep_tasks'].items():
        if tasks:
            print(f"- {category.title()}: {len(tasks)} tasks")
    
    print("Dynamic Reports generated successfully!")
    print("Files created:")
    print("- wo_analysis_chart.png")
    print("- ppm_chart.png") 
    print("- fm_report_template.html")
    print("- fm_report.xlsx")
    
    # Print the extracted metrics
    print(f"\nExtracted Metrics:")
    print(f"Date: {fm_report.report_data['date']}")
    print(f"Total WOs Raised: {fm_report.report_data['cm_wo_analysis']['total_raised']}")
    print(f"WOs Closed: {fm_report.report_data['cm_wo_analysis']['closed']}")
    print(f"WOs Pending: {fm_report.report_data['cm_wo_analysis']['pending']}")
    print(f"Total Work Orders in Summary: {len(fm_report.report_data['work_orders'])}")

# 6. MODIFY: Main execution
if __name__ == "__main__":
    # Usage with Excel file
    excel_file_path = "daily_report_2025-06-07.xlsx"  # Replace with actual file path
    pm_path="ppm_daily_report_2025-06-07_20250609_155958.xlsx"
    dynamic_example_usage(excel_file_path,pm_path)